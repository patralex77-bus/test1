import re
from openpyxl import load_workbook
from io import BytesIO
from datetime import date, datetime


def norm(s: str) -> str:
    return str(s or "").strip().lower().replace("\n", " ").replace("\t", " ")


def _row_is_empty(vals):
    return all(v is None or str(v).strip() == "" for v in vals)


def _strip_trailing_dot_zero(s: str) -> str:
    """
    '80965606896.0' -> '80965606896', '7433476.00' -> '7433476'
    """
    s = (s or "").strip()
    if not s:
        return s
    m = re.fullmatch(r"(\d+)\.0+", s)
    if m:
        return m.group(1)
    return s


def _to_str(val) -> str:
    if val is None:
        return ""
    # openpyxl може да върне float за телефони и т.н.
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val)
    if isinstance(val, (datetime, date)):
        # маркер за дата
        return "__DATE__"
    return str(val)


def _clean_text(s: str) -> str:
    s = (s or "").strip()
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s)
    return s


# ✅ НОВО: skip на "служебни" редове още преди мапването
def _is_separator_row(row_vals) -> bool:
    """
    Пропуска редове като:
      - "08" (ден от месеца като отделен ред)
      - "EUR"/"UAH" като отделен ред
      - "08" + "EUR" (две клетки)
      - Excel date/datetime като отделен ред
    """
    vals = []
    has_date = False

    for v in row_vals:
        s = _to_str(v)
        if s == "__DATE__":
            has_date = True
            continue
        s = _clean_text(_strip_trailing_dot_zero(s))
        if s:
            vals.append(s)

    if has_date and len(vals) == 0:
        return True

    # само 1 попълнена клетка
    if len(vals) == 1:
        x = vals[0].upper()
        if x in ("EUR", "UAH"):
            return True
        if re.fullmatch(r"\d{1,2}", vals[0]):  # "08" / "8"
            d = int(vals[0])
            if 1 <= d <= 31:
                return True
        # текстова дата
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", vals[0]) or re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", vals[0]):
            return True

    # 2 попълнени клетки: "08" + "EUR" или "EUR" + "08"
    if len(vals) == 2:
        a, b = vals[0], vals[1]
        au, bu = a.upper(), b.upper()
        if (au in ("EUR", "UAH") and re.fullmatch(r"\d{1,2}", b)) or (bu in ("EUR", "UAH") and re.fullmatch(r"\d{1,2}", a)):
            return True

    return False


def detect_currency(raw: str) -> str | None:
    s = (raw or "").lower()
    if re.search(r"(грн|uah|₴|гривн(?:а|и)?)", s, re.I):
        return "UAH"
    if re.search(r"(євро|евро|eur|€)", s, re.I):
        return "EUR"
    return None


PAID_WORDS_RE = re.compile(r"\b(сплачено|оплачено|сплачена|сплачений|paid)\b", re.I)


def parse_amount_from_text(raw: str):
    """
    Вади сума САМО ако има валута (както каза: винаги е обозначена).
    Примери:
      "160 євро" -> 160
      "4700 грн" -> 4700
      "€12.00"   -> 12
    Всички други числа без валута -> None
    """
    s = _clean_text(raw)
    if not s:
        return None

    s = _strip_trailing_dot_zero(s)

    has_currency = bool(re.search(r"(€|eur|євро|евро|uah|₴|грн|гривн(?:а|и)?)", s, re.I))
    if not has_currency:
        return None

    cleaned = re.sub(r"[^\d,.\-]", "", s).replace(",", ".")
    if not re.fullmatch(r"-?\d+(\.\d{1,2})?", cleaned):
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _is_day_marker_row(item: dict) -> bool:
    """
    Оставяме го като втори "safety net" след мапването,
    но основното skip вече става с _is_separator_row().
    """
    pn = (item.get("passenger_no") or "").strip()
    if not re.fullmatch(r"\d{1,2}", pn):
        return False
    day = int(pn)
    if not (1 <= day <= 31):
        return False

    from_city = _clean_text(item.get("from_city") or "")
    to_city = _clean_text(item.get("to_city") or "")
    full_name = _clean_text(item.get("full_name") or "")
    phone = _clean_text(item.get("phone") or "")
    seat = _clean_text(item.get("seat_no") or "")
    raw = _clean_text(item.get("voucher_or_amount_raw") or "")

    if full_name.upper() in ("EUR", "UAH"):
        full_name = ""

    if (
        from_city == "" and to_city == "" and phone == "" and seat == "" and full_name == ""
        and (raw == "" or raw.upper() in ("EUR", "UAH"))
    ):
        return True

    return False


HEADER_MAP = {
    "номер пасажир": "passenger_no",
    "київ - град начало": "from_city",
    "град начало": "from_city",
    "грац - град край": "to_city",
    "град край": "to_city",
    "име": "full_name",
    "номер място": "seat_no",
    "място": "seat_no",
    "телефон": "phone",
    "номер на ваучер или сума за плащане": "voucher_or_amount_raw",
    "ваучер или сума за плащане": "voucher_or_amount_raw",
    "ваучер/сума": "voucher_or_amount_raw",
}

# ✅ поправен ред на колоните за positional mode:
# №, От, До, Име, Телефон, Място, Инфо
POS_KEYS = [
    "passenger_no",
    "from_city",
    "to_city",
    "full_name",
    "phone",
    "seat_no",
    "voucher_or_amount_raw",
]


def _normalize_passenger_no(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return None

    s = _clean_text(str(val))
    if s.isdigit():
        return str(int(s))
    m = re.fullmatch(r"(\d+)\.0+", s)
    if m:
        return str(int(m.group(1)))
    return None


def _normalize_phone(val) -> str | None:
    s = _clean_text(_to_str(val))
    if not s:
        return None
    s = _strip_trailing_dot_zero(s)
    s = s.replace(" ", "")
    return s or None


def _postprocess_voucher_amount(item: dict):
    # phone
    if "phone" in item:
        item["phone"] = _normalize_phone(item["phone"])

    raw = _clean_text(item.get("voucher_or_amount_raw") or "")
    raw = _strip_trailing_dot_zero(raw)

    # винаги пазим RAW (колона Инфо)
    item["voucher_or_amount_raw"] = raw

    if not raw:
        item.setdefault("currency", "EUR")
        item["amount_due"] = None
        item["voucher_code"] = None
        return item

    # 1) "сплачено" => няма дължима сума
    if PAID_WORDS_RE.search(raw):
        item.setdefault("currency", "EUR")
        item["amount_due"] = None
        item["voucher_code"] = None
        return item

    # 2) валута от текста (EUR/UAH)
    cur = detect_currency(raw) or "EUR"
    item["currency"] = cur

    # 3) сума? (само ако има валута)
    money = parse_amount_from_text(raw)
    if money is not None:
        item["amount_due"] = money
        item["voucher_code"] = None
        return item

    # 4) ако е "само цифри" и >= 6 -> voucher_code
    only_digits = bool(re.fullmatch(r"\d+", raw))
    if only_digits and len(raw) >= 6:
        item["voucher_code"] = raw
        item["amount_due"] = None
        return item

    # 5) всичко друго -> код
    item["voucher_code"] = raw
    item["amount_due"] = None
    return item


def parse_xlsx(file_bytes: bytes):
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    best_idx = None
    best_score = 0
    for i in range(min(10, len(rows))):
        r = rows[i]
        if _row_is_empty(r) or _is_separator_row(r):
            continue
        headers = [norm(h) for h in r]
        score = sum(1 for h in headers if h in HEADER_MAP)
        if score > best_score:
            best_score = score
            best_idx = i

    out = []

    # A) headers mode
    if best_score >= 2 and best_idx is not None:
        headers = [norm(h) for h in rows[best_idx]]
        mapped = [HEADER_MAP.get(h) for h in headers]

        data_rows = rows[best_idx + 1 :]
        for r in data_rows:
            if _row_is_empty(r) or _is_separator_row(r):
                continue

            item = {}
            for idx, val in enumerate(r):
                key = mapped[idx] if idx < len(mapped) else None
                if not key:
                    continue
                item[key] = _clean_text(_to_str(val))

            if "passenger_no" in item:
                item["passenger_no"] = _normalize_passenger_no(item["passenger_no"])

            # ✅ втори safety net
            if item.get("passenger_no") and _is_day_marker_row(item):
                continue

            # ако няма име (или е само "EUR"/"UAH") => skip
            full = _clean_text(item.get("full_name") or "")
            if full.upper() in ("EUR", "UAH"):
                full = ""
            if not full:
                continue
            item["full_name"] = full

            out.append(_postprocess_voucher_amount(item))

        return out

    # B) positional mode
    for r in rows:
        if _row_is_empty(r) or _is_separator_row(r):
            continue

        pn = _normalize_passenger_no(r[0] if len(r) > 0 else None)
        if pn is None:
            continue

        item = {"passenger_no": pn}

        for idx, key in enumerate(POS_KEYS[1:], start=1):
            if idx < len(r):
                item[key] = _clean_text(_to_str(r[idx]))

        # ✅ safety net
        if _is_day_marker_row(item):
            continue

        full = _clean_text(item.get("full_name") or "")
        if full.upper() in ("EUR", "UAH"):
            full = ""
        if not full:
            continue
        item["full_name"] = full

        out.append(_postprocess_voucher_amount(item))

    return out
