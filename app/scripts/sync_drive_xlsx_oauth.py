import os
import io
import re
import json
import hashlib
import base64

from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, Tuple, Any

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google.oauth2.credentials import Credentials
from openpyxl import load_workbook

from app.db import SessionLocal
from app import crud

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

HERE = Path(__file__).resolve().parent
TOKEN_FILE = Path(os.environ.get("GOOGLE_OAUTH_TOKEN_FILE", str(HERE / "token.json")))
STATE_FILE = Path(os.environ.get("SYNC_STATE_FILE", str(HERE / "sync_state.json")))

DRIVE_FILE_ID = os.environ.get("DRIVE_FILE_ID", "").strip()

# Режими:
#   all     -> синхронизира всички sheet-ове
#   future  -> само днес и бъдеще
#   rolling -> последните N дни + бъдеще
SYNC_MODE = os.environ.get("SYNC_MODE", "all").strip().lower()
SYNC_DAYS_BACK = int(os.environ.get("SYNC_DAYS_BACK", "2"))


def ensure_token_file():
    b64 = os.environ.get("GOOGLE_TOKEN_JSON_B64", "").strip()
    if not b64:
        return

    raw = base64.b64decode(b64.encode("utf-8"))
    data = json.loads(raw.decode("utf-8"))

    if not data.get("refresh_token"):
        raise RuntimeError("GOOGLE_TOKEN_JSON_B64 does not contain refresh_token")

    TOKEN_FILE.parent.mkdir(parents=True, exist_ok=True)
    TOKEN_FILE.write_text(json.dumps(data), encoding="utf-8")


def download_drive_file_bytes(file_id: str) -> bytes:
    ensure_token_file()
    creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    service = build("drive", "v3", credentials=creds)

    request = service.files().export_media(
        fileId=file_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        _, done = downloader.next_chunk()

    return fh.getvalue()


def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def sha256_text(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def default_sync_state() -> dict[str, Any]:
    return {
        "version": 2,
        "file_hash": None,
        "trips": {},
        "updated_at": None,
    }


def read_sync_state() -> dict[str, Any]:
    if not STATE_FILE.exists():
        return default_sync_state()

    try:
        raw = STATE_FILE.read_text(encoding="utf-8").strip()
    except Exception:
        return default_sync_state()

    if not raw:
        return default_sync_state()

    # backward compatibility:
    # ако старият файл е бил просто plain hash string
    try:
        data = json.loads(raw)
    except Exception:
        st = default_sync_state()
        st["file_hash"] = raw
        return st

    if not isinstance(data, dict):
        return default_sync_state()

    st = default_sync_state()
    st["version"] = data.get("version", 2)
    st["file_hash"] = data.get("file_hash")
    st["trips"] = data.get("trips", {}) or {}
    st["updated_at"] = data.get("updated_at")
    return st


def write_sync_state(state: dict[str, Any]) -> None:
    try:
        STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
        state["updated_at"] = datetime.utcnow().isoformat()
        STATE_FILE.write_text(
            json.dumps(state, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8",
        )
    except Exception:
        pass


def _clean_text(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def _strip_trailing_dot_zero(s: str) -> str:
    s = (s or "").strip()
    m = re.fullmatch(r"(\d+)\.0+", s)
    return m.group(1) if m else s


def _normalize_phone(v) -> str:
    s = _clean_text(v)
    s = _strip_trailing_dot_zero(s)
    s = s.replace(" ", "")
    return s


def _normalize_uid(v) -> str:
    s = _clean_text(v)
    s = _strip_trailing_dot_zero(s)
    return s


def guess_trip_from_sheet(sheet_name: str) -> Tuple[Optional[date], Optional[str], Optional[str]]:
    s = (sheet_name or "").strip()

    parts = s.split()
    if len(parts) >= 2:
        try:
            date_only = datetime.strptime(parts[0], "%Y-%m-%d").date()
            route = parts[1]
            if "-" in route:
                a, b = route.split("-", 1)
                return date_only, a.strip(), b.strip()
        except Exception:
            pass

    m = re.match(r"^\s*(\d{1,2})\.(\d{1,2})\s*(KI|IK|КІ|ІК)\s*$", s, re.IGNORECASE)
    if not m:
        return None, None, None

    dd = int(m.group(1))
    mm = int(m.group(2))
    code_raw = m.group(3)

    code = code_raw.upper()
    code = code.replace("К", "K").replace("І", "I")

    year = int(os.environ.get("SYNC_YEAR", str(datetime.now().year)))

    try:
        date_only = datetime(year, mm, dd).date()
    except Exception:
        return None, None, None

    if code == "KI":
        return date_only, "Kyiv", "Innsbruck"
    if code == "IK":
        return date_only, "Innsbruck", "Kyiv"

    return None, None, None


def _is_day_marker_row(
    passenger_no: str,
    from_city: str,
    to_city: str,
    full_name: str,
    seat_no: str,
    phone: str,
    raw: str,
) -> bool:
    if not passenger_no.isdigit():
        return False

    d = int(passenger_no)
    if not (1 <= d <= 31):
        return False

    fn = (full_name or "").strip()
    if fn.upper() in ("EUR", "UAH"):
        fn = ""

    raw2 = (raw or "").strip()
    if raw2.upper() in ("EUR", "UAH"):
        raw2 = ""

    if (
        (from_city or "").strip() == ""
        and (to_city or "").strip() == ""
        and (seat_no or "").strip() == ""
        and (phone or "").strip() == ""
        and fn == ""
        and raw2 == ""
    ):
        return True

    return False


def _normalize_info_raw(v) -> str:
    if v is None:
        return ""

    if isinstance(v, (int, float)):
        try:
            n = float(v)
        except Exception:
            return _clean_text(v)

        if n <= 0 or n > 2000:
            return _clean_text(v)

        whole = int(abs(n))
        if whole > 9999:
            return _clean_text(v)

        return f"{n:.2f}"

    s = _clean_text(v)
    if not s:
        return ""

    m = re.fullmatch(r"(\d{1,4})([.,](\d{1,2}))?", s)
    if not m:
        return s

    try:
        n = float(s.replace(",", "."))
    except Exception:
        return s

    if n <= 0 or n > 2000:
        return s

    return f"{n:.2f}"


def parse_sheet_rows(wb, sheet_name: str, skip_rows: int = 1) -> list[dict]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    out: list[dict] = []

    def norm(v) -> str:
        return ("" if v is None else str(v)).strip()

    def is_currency_only(s: str) -> bool:
        return s.upper() in ("EUR", "UAH")

    for idx, r in enumerate(rows):
        if idx < skip_rows:
            continue
        if not r:
            continue

        vals = list(r) + [None] * 8

        a = vals[0]
        if a is None:
            continue

        s0 = norm(a)
        if not s0:
            continue

        s0_clean = s0.split(".")[0].strip()
        if not s0_clean.isdigit():
            continue

        from_city = norm(vals[1])
        to_city = norm(vals[2])
        full_name = norm(vals[3])
        seat_no = norm(vals[4])
        phone = _normalize_phone(vals[5])
        raw = _normalize_info_raw(vals[6])
        source_uid = _normalize_uid(vals[7])

        full_for_check = "" if is_currency_only(full_name) else full_name

        if (not full_for_check) and (not phone) and (not from_city) and (not to_city):
            continue

        if _is_day_marker_row(
            passenger_no=s0_clean,
            from_city=from_city,
            to_city=to_city,
            full_name=full_name,
            seat_no=seat_no,
            phone=phone,
            raw=raw,
        ):
            continue

        if is_currency_only(full_name):
            full_name = ""

        if not full_name:
            continue

        out.append(
            {
                "passenger_no": s0_clean,
                "from_city": from_city,
                "to_city": to_city,
                "full_name": full_name,
                "seat_no": seat_no,
                "phone": phone,
                "voucher_or_amount_raw": raw,
                "source_uid": source_uid or None,
            }
        )

    return out


def build_trip_key(date_only: date, route_from: str, route_to: str) -> str:
    return f"{date_only.isoformat()}|{route_from.strip()}|{route_to.strip()}"


def compute_trip_hash(date_only: date, route_from: str, route_to: str, rows: list[dict]) -> str:
    payload = {
        "date": date_only.isoformat(),
        "route_from": route_from,
        "route_to": route_to,
        "rows": rows,
    }
    return sha256_text(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    )


def should_sync_date(date_only: date, today: date) -> tuple[bool, str]:
    mode = SYNC_MODE

    if mode == "all":
        return True, ""

    if mode == "future":
        if date_only < today:
            return False, f"past date {date_only} < {today}"
        return True, ""

    if mode == "rolling":
        cutoff = today - timedelta(days=max(SYNC_DAYS_BACK, 0))
        if date_only < cutoff:
            return False, f"date {date_only} < rolling cutoff {cutoff}"
        return True, ""

    return True, ""


def find_or_create_trip(db, route_from: str, route_to: str, date_only, note: str):
    trip = crud.find_trip_by_route_date(db, route_from, route_to, date_only)
    if trip:
        if note and trip.note != note:
            trip.note = note
            db.commit()
            db.refresh(trip)
        return trip

    dt = datetime(date_only.year, date_only.month, date_only.day, 0, 0, 0)
    return crud.create_trip(db, route_from=route_from, route_to=route_to, date_time=dt, note=note)


def run_sync(force: bool = False) -> dict:
    if not DRIVE_FILE_ID:
        raise RuntimeError("Missing DRIVE_FILE_ID env var")
    if not TOKEN_FILE.exists() and not os.environ.get("GOOGLE_TOKEN_JSON_B64", "").strip():
        raise RuntimeError(f"Missing token file: {TOKEN_FILE}")

    today = datetime.now().date()
    state = read_sync_state()

    xlsx_bytes = download_drive_file_bytes(DRIVE_FILE_ID)
    file_hash = sha256_bytes(xlsx_bytes)

    if not force and state.get("file_hash") == file_hash:
        print("[SYNC] File hash unchanged. Skip all.")
        return {
            "ok": True,
            "changed": False,
            "imported": 0,
            "trips": 0,
            "cutoff": str(today),
            "mode": SYNC_MODE,
            "skipped_unchanged_trips": 0,
        }

    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
    sheetnames = wb.sheetnames

    imported_total = 0
    trips_touched = 0
    skipped_bad_name = 0
    skipped_by_policy = 0
    skipped_unchanged_trips = 0
    duplicate_trip_keys = 0

    prev_trips_state: dict[str, Any] = state.get("trips", {}) or {}
    new_trips_state: dict[str, Any] = dict(prev_trips_state)
    seen_trip_keys: set[str] = set()

    db = SessionLocal()
    try:
        for sh in sheetnames:
            date_only, rf, rt = guess_trip_from_sheet(sh)
            if not (date_only and rf and rt):
                skipped_bad_name += 1
                print(f"[SYNC] Skip sheet '{sh}' (bad name).")
                continue

            allowed, reason = should_sync_date(date_only, today)
            if not allowed:
                skipped_by_policy += 1
                print(f"[SYNC] Skip sheet '{sh}' ({reason}).")
                continue

            trip_key = build_trip_key(date_only, rf, rt)
            if trip_key in seen_trip_keys:
                duplicate_trip_keys += 1
                print(f"[SYNC] Skip sheet '{sh}' (duplicate trip key {trip_key}).")
                continue

            seen_trip_keys.add(trip_key)

            rows = parse_sheet_rows(wb, sh, skip_rows=1)
            trip_hash = compute_trip_hash(date_only, rf, rt, rows)
            prev = prev_trips_state.get(trip_key, {})

            if not force and prev.get("trip_hash") == trip_hash:
                skipped_unchanged_trips += 1
                print(f"[SYNC] Skip trip {trip_key} (unchanged).")
                new_trips_state[trip_key] = {
                    **prev,
                    "sheet_name": sh,
                    "date": date_only.isoformat(),
                    "route_from": rf,
                    "route_to": rt,
                    "rows_count": len(rows),
                }
                continue

            trip = find_or_create_trip(db, rf, rt, date_only, note=f"Auto: {sh}")
            imported = crud.import_passengers(db, trip.id, rows, replace=True)

            imported_total += int(imported or 0)
            trips_touched += 1

            new_trips_state[trip_key] = {
                "trip_hash": trip_hash,
                "sheet_name": sh,
                "date": date_only.isoformat(),
                "route_from": rf,
                "route_to": rt,
                "trip_id": trip.id,
                "rows_count": len(rows),
                "synced_at": datetime.utcnow().isoformat(),
            }

            print(
                f"[SYNC] {date_only} {rf}->{rt} trip_id={trip.id} "
                f"rows={len(rows)} synced={imported}"
            )

        state["version"] = 2
        state["file_hash"] = file_hash
        state["trips"] = new_trips_state
        write_sync_state(state)

        print("[SYNC] Done.")
        return {
            "ok": True,
            "changed": True,
            "cutoff": str(today),
            "mode": SYNC_MODE,
            "imported": imported_total,
            "trips": trips_touched,
            "skipped_bad_name": skipped_bad_name,
            "skipped_by_policy": skipped_by_policy,
            "skipped_unchanged_trips": skipped_unchanged_trips,
            "duplicate_trip_keys": duplicate_trip_keys,
        }

    finally:
        db.close()


def main():
    force = os.environ.get("FORCE_SYNC", "").strip().lower() in ("1", "true", "yes", "y")
    run_sync(force=force)


if __name__ == "__main__":
    main()
