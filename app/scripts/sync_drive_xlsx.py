import os
import io
import hashlib
from datetime import datetime
from typing import Optional

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

from openpyxl import load_workbook

# --- ТУК импортни твоите DB/crud/parse функции ---
# Примерно:
from app.db import SessionLocal
from app import crud
from app.excel_import import parse_xlsx  # ако parse_xlsx може да приема bytes
# Ако parse_xlsx е за първа вкладка, ще направим parse_xlsx_sheet по-долу.

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

DRIVE_FILE_ID = os.environ.get("DRIVE_FILE_ID", "").strip()
SERVICE_ACCOUNT_JSON = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "").strip()

# По желание: пазим хеш на последния свален файл, за да не правим работа ако няма промяна
STATE_FILE = os.environ.get("SYNC_STATE_FILE", "sync_state.txt")


def download_drive_file_bytes(file_id: str) -> bytes:
    creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    service = build("drive", "v3", credentials=creds)

    # Google Sheet -> export as XLSX
    request = service.files().export_media(
        fileId=file_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()

    return fh.getvalue()



def file_hash(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def read_last_hash() -> Optional[str]:
    if not os.path.exists(STATE_FILE):
        return None
    try:
        return open(STATE_FILE, "r", encoding="utf-8").read().strip() or None
    except Exception:
        return None


def write_last_hash(h: str) -> None:
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            f.write(h)
    except Exception:
        pass


def guess_trip_meta(sheet_name: str):
    """
    Тук решаваш как от името на вкладката да разберем дата и направление.
    Примерни варианти:
      - "2026-01-28 Innsbruck-Kyiv"
      - "28.01.2026 Kyiv-Innsbruck"
      - или просто "Innsbruck-Kyiv" и датата е в първия ред на sheet-а.

    Ще направя пример: "YYYY-MM-DD Innsbruck-Kyiv"
    """
    s = sheet_name.strip()
    parts = s.split()
    date = None
    route = None

    if len(parts) >= 2:
        # date
        try:
            date = datetime.strptime(parts[0], "%Y-%m-%d").date()
        except Exception:
            date = None
        route = parts[1]
    else:
        route = s

    rf, rt = None, None
    if route and "-" in route:
        a, b = route.split("-", 1)
        rf, rt = a.strip(), b.strip()

    return date, rf, rt


def parse_sheet_rows(wb, sheet_name: str, skip_rows: int = 1) -> list[dict]:
    """
    Чете редове от sheet и връща списък от dict.
    ✅ skip_rows=1 => винаги прескача първия ред (заглавие: дата+направление)
    ✅ Допълнително: пропуска "служебни" редове без име/тел/от/до.
    Очаквани колони:
      A: passenger_no
      B: from_city
      C: to_city
      D: full_name
      E: seat_no
      F: phone
      G: voucher_or_amount_raw
    """
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    out: list[dict] = []

    def norm(v) -> str:
        return ("" if v is None else str(v)).strip()

    def is_currency_only(s: str) -> bool:
        return s.upper() in ("EUR", "UAH")

    for idx, r in enumerate(rows):
        # ✅ 1) прескочи първите skip_rows реда (по подразбиране 1)
        if idx < skip_rows:
            continue

        if not r:
            continue

        vals = list(r) + [None] * 7

        a = vals[0]
        if a is None:
            continue

        # A може да е 1, 1.0, "001", понякога Excel дава float
        s0 = norm(a)
        if not s0:
            continue

        # ако е число като "1.0" -> "1"
        s0_clean = s0.split(".")[0].strip()

        # ✅ редът е валиден само ако passenger_no е число
        if not s0_clean.isdigit():
            continue

        from_city = norm(vals[1])
        to_city   = norm(vals[2])
        full_name = norm(vals[3])
        seat_no   = norm(vals[4])
        phone     = norm(vals[5])
        raw       = norm(vals[6])

        # ✅ 2) пропускаме “служебни” редове:
        # - ако няма абсолютно никакви ключови данни (име/тел/от/до),
        # - или full_name е само "EUR/UAH"
        full_for_check = "" if is_currency_only(full_name) else full_name

        if (not full_for_check) and (not phone) and (not from_city) and (not to_city):
            # типичен "08" / празен ред / служебен marker
            continue

        if is_currency_only(full_name):
            full_name = ""

        # ✅ ако няма име, но има други полета - прецени:
        # за да не вкарваш боклук, по-добре skip без име
        if not full_name:
            continue

        out.append({
            "passenger_no": s0_clean,
            "from_city": from_city,
            "to_city": to_city,
            "full_name": full_name,
            "seat_no": seat_no,
            "phone": phone,
            "voucher_or_amount_raw": raw,
        })

    return out


def upsert_trip(db, route_from: str, route_to: str, date_only, note: str = ""):
    """
    Най-добре е да имаш уникален ключ (route_from, route_to, date).
    Ако нямаш – ще направим “търси по тези 3” и ако няма -> create.
    """
    trip = crud.find_trip_by_route_date(db, route_from, route_to, date_only)  # ще добавиш тази функция
    if trip:
        # update note / date_time ако трябва
        if note:
            trip.note = note
        # date_time държиш като datetime 00:00
        db.commit()
        db.refresh(trip)
        return trip

    dt = datetime(date_only.year, date_only.month, date_only.day, 0, 0, 0)
    return crud.create_trip(db, route_from=route_from, route_to=route_to, date_time=dt, note=note)


def main():
    if not DRIVE_FILE_ID or not SERVICE_ACCOUNT_JSON:
        raise SystemExit("Missing DRIVE_FILE_ID or GOOGLE_APPLICATION_CREDENTIALS")

    xlsx_bytes = download_drive_file_bytes(DRIVE_FILE_ID)
    h = file_hash(xlsx_bytes)

    last = read_last_hash()
    if last == h:
        print("[SYNC] No changes, skip.")
        return

    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
    sheetnames = wb.sheetnames

    db = SessionLocal()
    try:
        for sh in sheetnames:
            date_only, rf, rt = guess_trip_meta(sh)
            if not (date_only and rf and rt):
                print(f"[SYNC] Skip sheet '{sh}' (can't parse meta).")
                continue

            trip = upsert_trip(db, rf, rt, date_only, note=f"Auto-import: {sh}")

            rows = parse_sheet_rows(xlsx_bytes, sh)

            # replace=True -> няма дубли, винаги е актуално
            crud.import_passengers(db, trip.id, rows, replace=True)
            print(f"[SYNC] {date_only} {rf}->{rt} trip_id={trip.id} passengers={len(rows)}")

        write_last_hash(h)
        print("[SYNC] Done.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
